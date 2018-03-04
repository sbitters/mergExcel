# -*- coding: utf-8 -*-
# Author: EVB

import openpyxl as pyxl
from sys import exit
from argparse import ArgumentParser


MIT_license = """Copyright 2018 Sven T. Bitters (sven.bitters@gmail.com)

Permission is hereby granted, free of charge, to any person obtaining a copy of
this software and associated documentation files (the "Software"), to deal in
the Software without restriction, including without limitation the rights to
use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of
the Software, and to permit persons to whom the Software is furnished to do so,
subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS
FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR
COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER
IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN
CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""


def handle_input():
    parser = ArgumentParser()
    parser.add_argument("-i", "--input", default="", dest="input_wbs", nargs="*")
    parser.add_argument("-s", "--sheets", default="first", dest="input_sheets", nargs="+")
    parser.add_argument("-t", "--target_wb", default="", dest="output_wb")
    parser.add_argument("--target_sheet", default="", dest="output_sheet")
    parser.add_argument("--target_col_titles_row", default=1, dest="output_col_titles_row")
    parser.add_argument("-r", "--rows", default="all", dest="merge_rows", nargs="*")
    parser.add_argument("-c", "--columns", default="all", dest="merge_cols", nargs="*")
    parser.add_argument("--row_titles_col", dest="row_titles_col", default="A", type=str, nargs="+")
    parser.add_argument("--col_titles_row", dest="col_titles_row", default=1, type=int, nargs="+")
    parser.add_argument("--separator", dest="separator_char", default=" / ", type=str)
    parser.add_argument("-l", "--license", action="store_true", default=False, dest="license")

    parserargs = parser.parse_args()

    if parserargs.license:
        exit(MIT_license)

    if len(parserargs.input_wbs) < 2:
        exit("Please input at least 2 Excel workbooks!")

    input_sheets = make_list(parserargs.input_sheets, parserargs.input_wbs)

    row_titles_col = colindex_to_number(make_list(parserargs.row_titles_col, parserargs.input_wbs))
    col_titles_row = make_list(parserargs.col_titles_row, parserargs.input_wbs)

    merge_rows = make_list(parserargs.merge_rows, parserargs.input_wbs)
    merge_cols = make_list(parserargs.merge_cols, parserargs.input_wbs)

    # option for no titles

    return {"input_wbs": parserargs.input_wbs, "input_sheets": input_sheets,
            "output_wb": parserargs.output_wb, "output_sheet": parserargs.output_sheet,
            "output_col_titles_row": parserargs.output_col_titles_row,
            "merge_rows": merge_rows, "merge_cols": merge_cols,
            "row_titles_col": row_titles_col, "col_titles_row": col_titles_row,
            "separator_char": parserargs.separator_char}


def make_list(arg, input_wbs):
    if type(arg) == list:
        if len(arg) == 1:
            arg *= len(input_wbs)
    elif len([arg]) == 1:
        arg = [arg] * len(input_wbs)

    return arg


def colindex_to_number(colindex):
    h = []
    for letter in colindex:
        if len(letter) == 1:
            h.append(ord(letter) - ord("A") + 1)
        else:
            number = 0
            for character in letter:
                number = number * 26 + ord(character.upper()) - ord("A") + 1
            h.append(number)

    return h


def number_to_colindex(col_number):
    h = []
    for number in col_number:
        if number <= 26:
            h.append(chr(number+64))
        else:
            colindex = ""
            while True:
                if number > 26:
                    number, remainder = divmod(number - 1, 26)
                    colindex = chr(remainder + ord('A')) + colindex
                else:
                    h.append(chr(number + ord("A") - 1) + colindex)
                    break
    return h


def read_data(p_args, wb_num):
    workbook = p_args["input_wbs"][wb_num]
    wb = pyxl.load_workbook(workbook, data_only=True)

    sheet = p_args["input_sheets"][wb_num]
    if sheet == "first":
        ws = wb[wb.get_sheet_names()[0]]
    else:
        ws = wb[sheet]

    data_dict = {}
    if p_args["merge_cols"] != make_list("all", p_args["input_wbs"]) and p_args["merge_rows"] == make_list("all", p_args["input_wbs"]):
        colval_list = []

        colnums = find_cols(ws, p_args["col_titles_row"][wb_num], p_args["col_titles_row"][wb_num], p_args["merge_cols"])

        for col in colnums:
            cells_data = []
            for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cells_data.append(str(cell.value))

            colval_list.append(cells_data)
            data_dict[cells_data[0]] = col

        ws_data = {data[0]: data[1:] for data in colval_list}

    elif p_args["merge_rows"] != make_list("all", p_args["input_wbs"]) and p_args["merge_cols"] == make_list("all", p_args["input_wbs"]):
        rowval_list = []

        rownums = find_rows(ws, p_args["row_titles_col"][wb_num], p_args["row_titles_col"][wb_num], p_args["merge_rows"])

        for row in rownums:
            cells_data = []
            for col in ws.iter_cols(min_row=row, max_row=row, min_col=p_args["row_titles_col"][wb_num]):
                for cell in col:
                    cells_data.append(str(cell.value))

            rowval_list.append(cells_data)
            data_dict[cells_data[0]] = row

        ws_data = {data[0]: data[1:] for data in rowval_list}

    elif p_args["merge_cols"] != make_list("all", p_args["input_wbs"]) and p_args["merge_rows"] != make_list("all", p_args["input_wbs"]):
        pass

    else:
        colval_list = []
        for col in ws.iter_cols(min_row=p_args["col_titles_row"][wb_num], min_col=p_args["row_titles_col"][wb_num]):
            vals = []
            for cell in col:
                vals.append(str(cell.value))

            colval_list.append(vals)

        ws_data = {mylist[0]: mylist[1:] for mylist in colval_list}

    return ws_data


def find_cols(ws, min_row, max_row, merge_cols):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        col_titles = [cell.value for cell in row]

        if not all(col_titles) is None:
            break

    colnums = set()
    for title in merge_cols:
        colnums.add(col_titles.index(title) + 1)

    return colnums


def find_rows(ws, min_col, max_col, merge_rows):
    for col in ws.iter_cols(min_col=min_col, max_col=max_col):
        row_titles = [cell.value for cell in col]

        if not all(row_titles) is None:
            break

    rownums = set()
    for title in merge_rows:
        rownums.add(row_titles.index(title) + 1)

    return rownums


def merge_cells(wb_data, p_args):

    merge_result = dict()

    for ws_set in wb_data:
        data_relevant = ws_set
        keys = [key for key in data_relevant.keys()]

        for key in keys:
            data_range = data_relevant[key]

            try:
                for ii in range(0, len(data_range)):
                    merge_result[key][ii] = merge_result[key][ii] + p_args["separator_char"] + data_range[ii]

            except KeyError:
                merge_result[key] = data_range

    return merge_result


def create_wb(merge_result, p_args):

    if p_args["output_wb"] and p_args["output_sheet"]:
        wb_out_name = p_args["output_wb"]
        wb_out = pyxl.load_workbook(filename=wb_out_name)
        ws_out = wb_out[p_args["output_sheet"]]

        if int(p_args["output_col_titles_row"]) < ws_out.min_row:
            min_row = ws_out.min_row
        else:
            min_row = int(p_args["output_col_titles_row"])

        colnums = find_cols(ws_out, min_row, ws_out.max_row, p_args["merge_cols"])

        colindex = number_to_colindex(colnums)

        for col in colindex:
            ii = -1
            for rownum in range(min_row, ws_out.max_row + 1):
                try:
                    if ii == -1:
                        data_to_insert = merge_result[(ws_out[col + str(rownum)]).value]

                    else:
                        ws_out[col + str(rownum)] = data_to_insert[ii]

                    ii += 1

                except IndexError:
                    break

        # rows

    else:
        wb_out_name = "OUTPUT_MERGE.xlsx"
        wb_out = pyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Merge Result"

        keys = sorted(list(merge_result.keys()))

        key_iter = 0
        for col in range(1, len(merge_result.keys())+1):
            col = number_to_colindex([col])[0]

            val_iter = 0
            for row in range(1, len(list(merge_result.values())[0])):

                if row == 1:
                    ws_out[col + str(row)] = keys[key_iter]
                else:
                    ws_out[col + str(row)] = merge_result[keys[key_iter]][val_iter]
                    val_iter += 1

            key_iter += 1

    wb_out.save(filename=wb_out_name)


def main():
    p_args = handle_input()

    wb_data = []
    for wb_num in range(0, len(p_args["input_wbs"])):
        wb_data.append(read_data(p_args, wb_num))

    merge_result = merge_cells(wb_data, p_args)

    create_wb(merge_result, p_args)


#######################################################################################

try:
	main()
	print("Done!")
except PermissionError:
	exit("Please close the output Excel workbook!")
